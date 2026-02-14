"""Employee Payroll Management System - GUI Application"""

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from datetime import datetime, date
import re, os
import database as db

# Optional imports
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

try:
    from PIL import Image, ImageDraw, ImageFont
    PIL_AVAILABLE = True
except ImportError:
    PIL_AVAILABLE = False

try:
    import matplotlib.pyplot as plt
    from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
    MATPLOTLIB_AVAILABLE = True
except ImportError:
    MATPLOTLIB_AVAILABLE = False


class PayrollApp:
    """Main application class for Employee Payroll Management System"""
    
    def __init__(self, root):
        self.root = root
        self.root.title("Employee Payroll Management System")
        self.root.geometry("1200x700")
        self.root.minsize(1000, 600)
        
        db.initialize_database()
        self.selected_photo_path = ""
        self.current_payroll = None
        
        self._setup_styles()
        self._create_menu()
        self._create_notebook()
        self._create_status_bar()
        self._refresh_employee_list()
        
    def _setup_styles(self):
        """Configure ttk styles"""
        style = ttk.Style()
        style.theme_use('clam')
        style.configure('TNotebook.Tab', padding=[20, 10], font=('Helvetica', 10))
        style.configure('TButton', padding=[10, 5], font=('Helvetica', 10))
        style.configure('Title.TLabel', font=('Helvetica', 14, 'bold'))
        style.configure('Treeview', rowheight=25)
        style.configure('Treeview.Heading', font=('Helvetica', 10, 'bold'))
        
    def _create_menu(self):
        """Create application menu bar"""
        menubar = tk.Menu(self.root)
        self.root.config(menu=menubar)
        
        # File menu
        file_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="File", menu=file_menu)
        file_menu.add_command(label="Export Employees to Excel", command=self._export_employees_to_excel)
        file_menu.add_command(label="Export Payroll to Excel", command=self._export_payroll_to_excel)
        file_menu.add_separator()
        file_menu.add_command(label="Exit", command=self.root.quit)
        
        # Tools menu
        tools_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="Tools", menu=tools_menu)
        tools_menu.add_command(label="Add Sample Employees", command=self._add_sample_employees)
        tools_menu.add_command(label="Attendance Analytics", command=self._show_attendance_analytics)
        
        # Help menu
        help_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="Help", menu=help_menu)
        help_menu.add_command(label="About", command=self._show_about)
        
    def _create_notebook(self):
        """Create tabbed interface"""
        self.notebook = ttk.Notebook(self.root)
        self.notebook.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
        self._create_employee_tab()
        self._create_attendance_tab()
        self._create_payroll_tab()
        
    def _create_status_bar(self):
        """Create status bar"""
        self.status_var = tk.StringVar(value="Ready")
        ttk.Label(self.root, textvariable=self.status_var, relief=tk.SUNKEN, 
                  anchor=tk.W).pack(side=tk.BOTTOM, fill=tk.X, padx=5, pady=2)
        
    def _set_status(self, message):
        """Update status bar"""
        self.status_var.set(message)
        self.root.update_idletasks()
    
    # Helper method for clearing treeview
    def _clear_tree(self, tree):
        for item in tree.get_children():
            tree.delete(item)
            
    # Validation helpers
    def _validate_email(self, email):
        if not email:
            return True
        return bool(re.match(r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$', email))
    
    def _validate_numeric(self, value, field_name, allow_negative=False):
        try:
            parsed = float(value)
            if not allow_negative and parsed < 0:
                return False, 0, f"{field_name} cannot be negative"
            return True, parsed, ""
        except ValueError:
            return False, 0, f"{field_name} must be a valid number"
    
    def _get_form_data(self):
        """Get and validate employee form data"""
        data = {k: self.emp_fields[k].get().strip() for k in self.emp_fields}
        data['tax_percent'] = data['tax_percent'] or "0"
        data['pf_enabled'] = self.pf_var.get()
        data['photo_path'] = self.selected_photo_path
        
        # Validate employee ID
        if not data['emp_id']:
            return None, "Employee ID is required"
        if len(data['emp_id']) > 20:
            return None, "Employee ID must be 20 characters or less"
        if not data['name']:
            return None, "Employee name is required"
            
        # Validate salary
        valid, data['basic_salary'], error = self._validate_numeric(data['basic_salary'], "Basic Salary")
        if not valid:
            return None, error
            
        # Validate tax
        valid, data['tax_percent'], error = self._validate_numeric(data['tax_percent'], "Tax Percent")
        if not valid:
            return None, error
        if data['tax_percent'] > 100:
            return None, "Tax percent cannot exceed 100%"
            
        # Validate email
        if not self._validate_email(data['email']):
            return None, "Invalid email format. Must contain @ and domain"
        return data, None
    
    # ==================== Employee Tab ====================
    
    def _create_employee_tab(self):
        """Create employee management tab"""
        self.emp_tab = ttk.Frame(self.notebook)
        self.notebook.add(self.emp_tab, text="Employee Management")
        
        # Left frame - form
        left_frame = ttk.LabelFrame(self.emp_tab, text="Employee Details", padding=10)
        left_frame.pack(side=tk.LEFT, fill=tk.Y, padx=5, pady=5)
        
        # Form fields
        self.emp_fields = {}
        fields = [("emp_id", "Employee ID*"), ("name", "Full Name*"), 
                  ("designation", "Designation"), ("department", "Department"),
                  ("basic_salary", "Basic Salary*"), ("bank_acc", "Bank Account"),
                  ("tax_percent", "Tax Percent (%)"), ("email", "Email"),
                  ("dob", "Date of Birth (YYYY-MM-DD)")]
        
        for i, (name, label) in enumerate(fields):
            ttk.Label(left_frame, text=label).grid(row=i, column=0, sticky=tk.W, pady=3)
            entry = ttk.Entry(left_frame, width=30)
            entry.grid(row=i, column=1, padx=5, pady=3)
            self.emp_fields[name] = entry
            
        # PF checkbox
        self.pf_var = tk.IntVar(value=1)
        ttk.Checkbutton(left_frame, text="PF Enabled", variable=self.pf_var).grid(
            row=len(fields), column=0, columnspan=2, sticky=tk.W, pady=5)
        
        # Photo selection
        photo_frame = ttk.Frame(left_frame)
        photo_frame.grid(row=len(fields)+1, column=0, columnspan=2, pady=5)
        ttk.Button(photo_frame, text="Select Photo", command=self._select_photo).pack(side=tk.LEFT, padx=5)
        self.photo_label = ttk.Label(photo_frame, text="No photo selected")
        self.photo_label.pack(side=tk.LEFT)
        
        # Buttons
        btn_frame = ttk.Frame(left_frame)
        btn_frame.grid(row=len(fields)+2, column=0, columnspan=2, pady=15)
        for text, cmd in [("Add Employee", self._add_employee), ("Update Employee", self._update_employee),
                          ("Delete Employee", self._delete_employee), ("Clear Form", self._clear_employee_form)]:
            ttk.Button(btn_frame, text=text, command=cmd).pack(side=tk.LEFT, padx=3)
        
        ttk.Button(left_frame, text="Generate ID Card", command=self._generate_id_card).grid(
            row=len(fields)+3, column=0, columnspan=2, pady=5)
        
        # Right frame - employee list
        right_frame = ttk.LabelFrame(self.emp_tab, text="Employee List", padding=10)
        right_frame.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        # Search
        search_frame = ttk.Frame(right_frame)
        search_frame.pack(fill=tk.X, pady=5)
        ttk.Label(search_frame, text="Search:").pack(side=tk.LEFT)
        self.search_var = tk.StringVar()
        self.search_var.trace('w', self._on_search)
        ttk.Entry(search_frame, textvariable=self.search_var, width=30).pack(side=tk.LEFT, padx=5)
        ttk.Button(search_frame, text="Refresh", command=self._refresh_employee_list).pack(side=tk.RIGHT)
        
        # Employee treeview
        columns = ("ID", "Name", "Designation", "Department", "Salary", "Email")
        self.emp_tree = ttk.Treeview(right_frame, columns=columns, show="headings", height=20)
        for col in columns:
            self.emp_tree.heading(col, text=col)
            self.emp_tree.column(col, width=100)
        self.emp_tree.column("Name", width=150)
        self.emp_tree.column("Email", width=180)
        
        scrollbar = ttk.Scrollbar(right_frame, orient=tk.VERTICAL, command=self.emp_tree.yview)
        self.emp_tree.configure(yscrollcommand=scrollbar.set)
        self.emp_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.emp_tree.bind('<<TreeviewSelect>>', self._on_employee_select)
        
    def _select_photo(self):
        """Open dialog to select photo"""
        path = filedialog.askopenfilename(title="Select Employee Photo",
            filetypes=[("Image Files", "*.png *.jpg *.jpeg *.gif *.bmp")])
        if path:
            self.selected_photo_path = path
            self.photo_label.config(text=os.path.basename(path))
            
    def _clear_employee_form(self):
        """Clear form fields"""
        for field in self.emp_fields.values():
            field.delete(0, tk.END)
        self.pf_var.set(1)
        self.selected_photo_path = ""
        self.photo_label.config(text="No photo selected")
        self.emp_tree.selection_remove(self.emp_tree.selection())
        
    def _refresh_employee_list(self):
        """Refresh employee list"""
        try:
            self._clear_tree(self.emp_tree)
            employees = db.get_all_employees()
            for emp in employees:
                self.emp_tree.insert('', tk.END, values=(
                    emp['emp_id'], emp['name'], emp['designation'],
                    emp['department'], f"${emp['basic_salary']:,.2f}", emp['email'] or ""))
            self._set_status(f"Loaded {len(employees)} employees")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load employees: {e}")
            
    def _on_search(self, *args):
        """Handle search"""
        try:
            self._clear_tree(self.emp_tree)
            term = self.search_var.get().strip()
            employees = db.search_employees(term) if term else db.get_all_employees()
            for emp in employees:
                self.emp_tree.insert('', tk.END, values=(
                    emp['emp_id'], emp['name'], emp['designation'],
                    emp['department'], f"${emp['basic_salary']:,.2f}", emp['email'] or ""))
        except Exception as e:
            self._set_status(f"Search error: {e}")
            
    def _on_employee_select(self, event):
        """Handle selection"""
        selection = self.emp_tree.selection()
        if not selection:
            return
        emp_id = self.emp_tree.item(selection[0])['values'][0]
        try:
            employee = db.get_employee(emp_id)
            if employee:
                self._populate_employee_form(employee)
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load employee: {e}")
            
    def _populate_employee_form(self, employee):
        """Fill form with employee data"""
        self._clear_employee_form()
        for field in self.emp_fields:
            if employee[field] is not None:
                self.emp_fields[field].insert(0, str(employee[field]))
        self.pf_var.set(employee['pf_enabled'])
        if employee['photo_path']:
            self.selected_photo_path = employee['photo_path']
            self.photo_label.config(text=os.path.basename(employee['photo_path']))
            
    def _add_employee(self):
        """Add new employee"""
        data, error = self._get_form_data()
        if error:
            messagebox.showerror("Validation Error", error)
            return
        try:
            db.add_employee(emp_id=data['emp_id'], name=data['name'], designation=data['designation'],
                department=data['department'], basic_salary=data['basic_salary'], bank_acc=data['bank_acc'],
                pf_enabled=data['pf_enabled'], tax_percent=data['tax_percent'], email=data['email'],
                dob=data['dob'], photo_path=data['photo_path'])
            messagebox.showinfo("Success", f"Employee '{data['name']}' added successfully!")
            self._clear_employee_form()
            self._refresh_employee_list()
            self._set_status(f"Added employee: {data['emp_id']}")
        except Exception as e:
            messagebox.showerror("Error", str(e))
            
    def _update_employee(self):
        """Update existing employee"""
        data, error = self._get_form_data()
        if error:
            messagebox.showerror("Validation Error", error)
            return
        if not db.employee_exists(data['emp_id']):
            messagebox.showerror("Error", f"Employee '{data['emp_id']}' does not exist")
            return
        try:
            db.update_employee(emp_id=data['emp_id'], name=data['name'], designation=data['designation'],
                department=data['department'], basic_salary=data['basic_salary'], bank_acc=data['bank_acc'],
                pf_enabled=data['pf_enabled'], tax_percent=data['tax_percent'], email=data['email'],
                dob=data['dob'], photo_path=data['photo_path'])
            messagebox.showinfo("Success", f"Employee '{data['name']}' updated successfully!")
            self._refresh_employee_list()
            self._set_status(f"Updated employee: {data['emp_id']}")
        except Exception as e:
            messagebox.showerror("Error", str(e))
            
    def _delete_employee(self):
        """Delete employee from database"""
        emp_id = self.emp_fields['emp_id'].get().strip()
        if not emp_id:
            messagebox.showerror("Error", "Please select an employee to delete")
            return
        if not messagebox.askyesno("Confirm Delete",
            f"Are you sure you want to delete employee '{emp_id}'?\nThis will also delete their attendance and payroll records."):
            return
        try:
            if db.delete_employee(emp_id):
                messagebox.showinfo("Success", f"Employee '{emp_id}' deleted successfully!")
                self._clear_employee_form()
                self._refresh_employee_list()
                self._set_status(f"Deleted employee: {emp_id}")
            else:
                messagebox.showerror("Error", f"Employee '{emp_id}' not found")
        except Exception as e:
            messagebox.showerror("Error", str(e))
            
    def _generate_id_card(self):
        """Generate ID card for selected employee"""
        if not PIL_AVAILABLE:
            messagebox.showerror("Error", "Pillow library required. Install: pip install Pillow")
            return
        emp_id = self.emp_fields['emp_id'].get().strip()
        if not emp_id:
            messagebox.showerror("Error", "Please select an employee")
            return
        employee = db.get_employee(emp_id)
        if not employee:
            messagebox.showerror("Error", f"Employee '{emp_id}' not found")
            return
        try:
            card_w, card_h = 400, 250
            card = Image.new('RGB', (card_w, card_h), 'white')
            draw = ImageDraw.Draw(card)
            draw.rectangle([5, 5, card_w-5, card_h-5], outline='#003366', width=3)
            draw.rectangle([5, 5, card_w-5, 50], fill='#003366')
            
            # Load fonts
            try:
                for path in ["arial.ttf", "C:/Windows/Fonts/arial.ttf"]:
                    try:
                        h_font = ImageFont.truetype(path, 20)
                        t_font = ImageFont.truetype(path, 14)
                        s_font = ImageFont.truetype(path, 12)
                        break
                    except: continue
                else: raise Exception()
            except: h_font = t_font = s_font = ImageFont.load_default()
                
            draw.text((card_w//2, 28), "EMPLOYEE ID CARD", fill='white', font=h_font, anchor='mm')
            
            # Photo
            px, py, ps = 30, 70, 80
            if employee['photo_path'] and os.path.exists(employee['photo_path']):
                try:
                    photo = Image.open(employee['photo_path']).resize((ps, ps))
                    card.paste(photo, (px, py))
                except:
                    draw.rectangle([px, py, px+ps, py+ps], outline='gray')
                    draw.text((px+ps//2, py+ps//2), "Photo", fill='gray', anchor='mm')
            else:
                draw.rectangle([px, py, px+ps, py+ps], outline='gray')
                draw.text((px+ps//2, py+ps//2), "Photo", fill='gray', anchor='mm')
                
            # Details
            tx = 130
            draw.text((tx, 75), f"ID: {employee['emp_id']}", fill='black', font=t_font)
            draw.text((tx, 100), f"Name: {employee['name']}", fill='black', font=t_font)
            draw.text((tx, 125), f"Dept: {employee['department'] or 'N/A'}", fill='black', font=t_font)
            draw.text((tx, 150), f"Designation: {employee['designation'] or 'N/A'}", fill='black', font=s_font)
            draw.text((card_w//2, 200), f"Valid until: {datetime.now().year + 1}-12-31", fill='gray', font=s_font, anchor='mm')
            draw.text((card_w//2, 220), "This card is property of the company", fill='gray', font=s_font, anchor='mm')
            
            path = filedialog.asksaveasfilename(defaultextension=".png", filetypes=[("PNG Image", "*.png")],
                initialfile=f"IDCard_{emp_id}.png")
            if path:
                card.save(path)
                messagebox.showinfo("Success", f"ID Card saved to:\n{path}")
                self._set_status(f"ID Card generated for {emp_id}")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to generate ID card: {e}")
            
    # ==================== Attendance Tab ====================
    
    def _create_attendance_tab(self):
        """Create attendance tab"""
        self.att_tab = ttk.Frame(self.notebook)
        self.notebook.add(self.att_tab, text="Attendance")
        
        # Mark attendance frame
        mark_frame = ttk.LabelFrame(self.att_tab, text="Mark Attendance", padding=10)
        mark_frame.pack(fill=tk.X, padx=5, pady=5)
        
        ttk.Label(mark_frame, text="Employee ID:").grid(row=0, column=0, padx=5, pady=5)
        self.att_emp_id = ttk.Entry(mark_frame, width=20)
        self.att_emp_id.grid(row=0, column=1, padx=5, pady=5)
        
        ttk.Label(mark_frame, text="Date (YYYY-MM-DD):").grid(row=0, column=2, padx=5, pady=5)
        self.att_date = ttk.Entry(mark_frame, width=15)
        self.att_date.grid(row=0, column=3, padx=5, pady=5)
        self.att_date.insert(0, date.today().strftime("%Y-%m-%d"))
        
        ttk.Label(mark_frame, text="Status:").grid(row=0, column=4, padx=5, pady=5)
        self.att_status = ttk.Combobox(mark_frame, width=12, values=["Present", "Absent", "Half-Day", "Leave"])
        self.att_status.grid(row=0, column=5, padx=5, pady=5)
        self.att_status.set("Present")
        ttk.Button(mark_frame, text="Mark Attendance", command=self._mark_attendance).grid(row=0, column=6, padx=10, pady=5)
            
        # Attendance list frame
        list_frame = ttk.LabelFrame(self.att_tab, text="Attendance Records", padding=10)
        list_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        filter_frame = ttk.Frame(list_frame)
        filter_frame.pack(fill=tk.X, pady=5)
        ttk.Label(filter_frame, text="Filter by Month:").pack(side=tk.LEFT)
        self.att_month_filter = ttk.Entry(filter_frame, width=10)
        self.att_month_filter.pack(side=tk.LEFT, padx=5)
        self.att_month_filter.insert(0, date.today().strftime("%Y-%m"))
        ttk.Button(filter_frame, text="Filter", command=self._refresh_attendance_list).pack(side=tk.LEFT, padx=5)
        ttk.Button(filter_frame, text="Show All", command=lambda: self._refresh_attendance_list(show_all=True)).pack(side=tk.LEFT, padx=5)
            
        # Attendance treeview
        columns = ("Date", "Employee ID", "Name", "Status")
        self.att_tree = ttk.Treeview(list_frame, columns=columns, show="headings", height=15)
        for col in columns:
            self.att_tree.heading(col, text=col)
            self.att_tree.column(col, width=150)
        scrollbar = ttk.Scrollbar(list_frame, orient=tk.VERTICAL, command=self.att_tree.yview)
        self.att_tree.configure(yscrollcommand=scrollbar.set)
        self.att_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self._refresh_attendance_list()
        
    def _mark_attendance(self):
        """Mark attendance for employee"""
        emp_id, att_date, status = self.att_emp_id.get().strip(), self.att_date.get().strip(), self.att_status.get()
        if not emp_id:
            messagebox.showerror("Error", "Please enter Employee ID")
            return
        if not att_date:
            messagebox.showerror("Error", "Please enter date")
            return
        if not db.employee_exists(emp_id):
            messagebox.showerror("Error", f"Employee ID '{emp_id}' does not exist.\nPlease add the employee first.")
            return
        try:
            db.mark_attendance(emp_id, att_date, status)
            messagebox.showinfo("Success", f"Attendance marked: {emp_id} - {status} on {att_date}")
            self._refresh_attendance_list()
            self._set_status(f"Attendance marked for {emp_id}")
        except Exception as e:
            messagebox.showerror("Error", str(e))
            
    def _refresh_attendance_list(self, show_all=False):
        """Refresh attendance list"""
        try:
            self._clear_tree(self.att_tree)
            month = None if show_all else self.att_month_filter.get().strip()
            records = db.get_all_attendance(month)
            for record in records:
                self.att_tree.insert('', tk.END, values=(record['date'], record['emp_id'], record['name'], record['status']))
            self._set_status(f"Loaded {len(records)} attendance records")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load attendance: {e}")
            
    # ==================== Payroll Tab ====================
    
    def _create_payroll_tab(self):
        """Create payroll tab"""
        self.pay_tab = ttk.Frame(self.notebook)
        self.notebook.add(self.pay_tab, text="Payroll")
        
        # Generate payroll frame
        gen_frame = ttk.LabelFrame(self.pay_tab, text="Generate Payroll", padding=10)
        gen_frame.pack(fill=tk.X, padx=5, pady=5)
        
        ttk.Label(gen_frame, text="Employee ID:").grid(row=0, column=0, padx=5, pady=5)
        self.pay_emp_id = ttk.Entry(gen_frame, width=20)
        self.pay_emp_id.grid(row=0, column=1, padx=5, pady=5)
        
        ttk.Label(gen_frame, text="Month (YYYY-MM):").grid(row=0, column=2, padx=5, pady=5)
        self.pay_month = ttk.Entry(gen_frame, width=12)
        self.pay_month.grid(row=0, column=3, padx=5, pady=5)
        self.pay_month.insert(0, date.today().strftime("%Y-%m"))
        
        ttk.Button(gen_frame, text="Calculate Payroll", command=self._calculate_payroll).grid(row=0, column=4, padx=10, pady=5)
        ttk.Button(gen_frame, text="Generate All", command=self._generate_all_payroll).grid(row=0, column=5, padx=5, pady=5)
            
        # Payroll details frame
        details_frame = ttk.LabelFrame(self.pay_tab, text="Payroll Breakdown", padding=10)
        details_frame.pack(fill=tk.X, padx=5, pady=5)
        
        self.payroll_labels = {}
        payroll_fields = [("Employee", "employee_name"), ("Basic Salary", "basic_salary"), ("HRA (20%)", "hra"),
            ("Gross Salary", "gross_salary"), ("PF Deduction (12%)", "pf_deduction"), ("Tax Deduction", "tax_deduction"),
            ("Leave Deduction", "leave_deduction"), ("Net Salary", "net_salary")]
        
        for i, (label, key) in enumerate(payroll_fields):
            row, col = divmod(i, 4)
            ttk.Label(details_frame, text=f"{label}:").grid(row=row, column=col*2, sticky=tk.E, padx=5, pady=3)
            lbl = ttk.Label(details_frame, text="-", width=15)
            lbl.grid(row=row, column=col*2+1, sticky=tk.W, padx=5, pady=3)
            self.payroll_labels[key] = lbl
            
        self.save_payroll_btn = ttk.Button(details_frame, text="Save Payroll", command=self._save_payroll, state=tk.DISABLED)
        self.save_payroll_btn.grid(row=2, column=6, columnspan=2, pady=10)
        
        # Payroll history
        history_frame = ttk.LabelFrame(self.pay_tab, text="Payroll History", padding=10)
        history_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        columns = ("Month", "Employee", "Gross", "PF", "Tax", "Leave Ded.", "Net Salary")
        self.pay_tree = ttk.Treeview(history_frame, columns=columns, show="headings", height=10)
        for col in columns:
            self.pay_tree.heading(col, text=col)
            self.pay_tree.column(col, width=100)
        scrollbar = ttk.Scrollbar(history_frame, orient=tk.VERTICAL, command=self.pay_tree.yview)
        self.pay_tree.configure(yscrollcommand=scrollbar.set)
        self.pay_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self._refresh_payroll_history()
        
    def _calculate_payroll(self):
        """Calculate payroll for employee"""
        emp_id, month = self.pay_emp_id.get().strip(), self.pay_month.get().strip()
        if not emp_id:
            messagebox.showerror("Error", "Please enter Employee ID")
            return
        if not month:
            messagebox.showerror("Error", "Please enter month")
            return
        try:
            payroll = db.calculate_payroll(emp_id, month)
            self.current_payroll = payroll
            # Update all labels using loop
            self.payroll_labels['employee_name'].config(text=payroll['employee_name'])
            for key in ['basic_salary', 'hra', 'gross_salary', 'pf_deduction', 'tax_deduction', 'leave_deduction', 'net_salary']:
                self.payroll_labels[key].config(text=f"${payroll[key]:,.2f}")
            self.save_payroll_btn.config(state=tk.NORMAL)
            self._set_status(f"Calculated payroll for {emp_id}")
        except Exception as e:
            messagebox.showerror("Error", str(e))
            
    def _save_payroll(self):
        """Save current payroll"""
        if not self.current_payroll:
            messagebox.showerror("Error", "No payroll to save. Calculate first.")
            return
        try:
            db.save_payroll(self.current_payroll)
            messagebox.showinfo("Success", "Payroll saved successfully!")
            self._refresh_payroll_history()
            self._set_status(f"Saved payroll for {self.current_payroll['emp_id']}")
        except Exception as e:
            messagebox.showerror("Error", str(e))
            
    def _generate_all_payroll(self):
        """Generate payroll for all employees"""
        month = self.pay_month.get().strip()
        if not month:
            messagebox.showerror("Error", "Please enter month")
            return
        if not messagebox.askyesno("Confirm", f"Generate payroll for all employees for {month}?"):
            return
        try:
            employees = db.get_all_employees()
            count = 0
            for emp in employees:
                try:
                    db.save_payroll(db.calculate_payroll(emp['emp_id'], month))
                    count += 1
                except: pass
            messagebox.showinfo("Success", f"Generated payroll for {count} employees")
            self._refresh_payroll_history()
            self._set_status(f"Generated {count} payrolls for {month}")
        except Exception as e:
            messagebox.showerror("Error", str(e))
            
    def _refresh_payroll_history(self):
        """Refresh payroll history"""
        try:
            self._clear_tree(self.pay_tree)
            for rec in db.get_payroll_history():
                self.pay_tree.insert('', tk.END, values=(rec['month'], rec['name'],
                    f"${rec['gross_salary']:,.2f}", f"${rec['pf_deduction']:,.2f}", f"${rec['tax_deduction']:,.2f}",
                    f"${rec['leave_deduction']:,.2f}", f"${rec['net_salary']:,.2f}"))
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load payroll history: {e}")
            
    # ==================== Excel Export (helper to reduce duplication) ====================
    
    def _export_to_excel(self, data, headers, field_map, title, filename):
        """Generic Excel export helper"""
        if not EXCEL_AVAILABLE:
            messagebox.showerror("Error", "openpyxl required. Install: pip install openpyxl")
            return
        if not data:
            messagebox.showinfo("Info", f"No {title.lower()} to export")
            return
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = title
            
            header_fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.fill, cell.font, cell.alignment = header_fill, header_font, Alignment(horizontal="center")
                
            for row_num, item in enumerate(data, 2):
                for col_num, field in enumerate(field_map, 1):
                    val = field(item) if callable(field) else item.get(field, "")
                    ws.cell(row=row_num, column=col_num, value=val)
                    
            for col in ws.columns:
                ws.column_dimensions[col[0].column_letter].width = max(len(str(c.value or "")) for c in col) + 2
                
            path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel Files", "*.xlsx")], initialfilename=filename)
            if path:
                wb.save(path)
                messagebox.showinfo("Success", f"Exported to:\n{path}")
                self._set_status(f"{title} exported to Excel")
        except Exception as e:
            messagebox.showerror("Error", f"Export failed: {e}")
    
    def _export_employees_to_excel(self):
        """Export employees to Excel"""
        headers = ["Employee ID", "Name", "Designation", "Department", "Basic Salary", "Bank Account", "PF Enabled", "Tax %", "Email", "DOB"]
        fields = ['emp_id', 'name', 'designation', 'department', 'basic_salary', 'bank_acc', 
                  lambda e: "Yes" if e['pf_enabled'] else "No", 'tax_percent', 'email', 'dob']
        self._export_to_excel(db.get_all_employees(), headers, fields, "Employees", "Employees.xlsx")
            
    def _export_payroll_to_excel(self):
        """Export payroll to Excel"""
        headers = ["Month", "Employee ID", "Name", "Basic", "HRA", "Gross", "PF Ded.", "Tax Ded.", "Leave Ded.", "Net Salary", "Generated On"]
        fields = ['month', 'emp_id', 'name', 'basic_salary', 'hra', 'gross_salary', 'pf_deduction', 'tax_deduction', 'leave_deduction', 'net_salary', 'generated_on']
        self._export_to_excel(db.get_payroll_history(), headers, fields, "Payroll", "Payroll.xlsx")
            
    def _add_sample_employees(self):
        """Add sample employees"""
        if not messagebox.askyesno("Confirm", "Add sample employees for testing?"):
            return
        try:
            count = db.bulk_add_sample_employees()
            messagebox.showinfo("Success", f"Added {count} sample employees")
            self._refresh_employee_list()
            self._set_status(f"Added {count} sample employees")
        except Exception as e:
            messagebox.showerror("Error", str(e))
            
    def _show_attendance_analytics(self):
        """Show attendance analytics with pie chart"""
        if not MATPLOTLIB_AVAILABLE:
            messagebox.showerror("Error", "matplotlib required. Install: pip install matplotlib")
            return
            
        win = tk.Toplevel(self.root)
        win.title("Attendance Analytics")
        win.geometry("800x500")
        
        controls = ttk.Frame(win, padding=10)
        controls.pack(fill=tk.X)
        
        ttk.Label(controls, text="Employee ID:").pack(side=tk.LEFT)
        emp_entry = ttk.Entry(controls, width=15)
        emp_entry.pack(side=tk.LEFT, padx=5)
        
        ttk.Label(controls, text="Month (YYYY-MM):").pack(side=tk.LEFT, padx=5)
        month_entry = ttk.Entry(controls, width=10)
        month_entry.pack(side=tk.LEFT, padx=5)
        month_entry.insert(0, date.today().strftime("%Y-%m"))
        
        chart_frame = ttk.Frame(win)
        chart_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        def generate_chart():
            emp_id, month = emp_entry.get().strip(), month_entry.get().strip()
            if not emp_id:
                messagebox.showerror("Error", "Please enter Employee ID")
                return
            try:
                summary = db.get_attendance_summary(emp_id, month)
                employee = db.get_employee(emp_id)
                if not employee:
                    messagebox.showerror("Error", f"Employee '{emp_id}' not found")
                    return
                    
                for widget in chart_frame.winfo_children():
                    widget.destroy()
                    
                fig, ax = plt.subplots(figsize=(8, 5))
                labels = ['Present', 'Absent', 'Half-Day', 'Leave']
                sizes = [summary['present'], summary['absent'], summary['half_day'], summary['leave']]
                colors = ['#2ecc71', '#e74c3c', '#f39c12', '#3498db']
                
                # Filter non-zero values
                data = [(l, s, c) for l, s, c in zip(labels, sizes, colors) if s > 0]
                if data:
                    labels, sizes, colors = zip(*data)
                    ax.pie(sizes, explode=[0.05]*len(sizes), labels=labels, colors=colors, autopct='%1.1f%%', shadow=True, startangle=90)
                    ax.axis('equal')
                    ax.set_title(f"Attendance for {employee['name']} - {month}")
                else:
                    ax.text(0.5, 0.5, "No attendance data", ha='center', va='center', fontsize=14)
                    ax.set_title(f"No data for {month}")
                    
                canvas = FigureCanvasTkAgg(fig, master=chart_frame)
                canvas.draw()
                canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)
            except Exception as e:
                messagebox.showerror("Error", str(e))
                
        ttk.Button(controls, text="Generate Chart", command=generate_chart).pack(side=tk.LEFT, padx=10)
                   
    def _show_about(self):
        """Show about dialog"""
        messagebox.showinfo("About", "Employee Payroll Management System\n\nVersion 1.0\n\n"
            "Features:\n- Employee Management\n- Attendance Tracking\n- Payroll Calculation\n"
            "- Excel Export\n- ID Card Generation\n- Attendance Analytics")


def main():
    root = tk.Tk()
    PayrollApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()
